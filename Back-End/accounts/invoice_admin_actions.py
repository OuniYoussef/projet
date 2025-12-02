"""
Actions d'export pour les factures (PDF et Excel)
"""
from django.http import HttpResponse
from django.utils.dateformat import format as date_format
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def export_invoices_to_excel(modeladmin, request, queryset):
    """
    Action pour exporter les factures sélectionnées en Excel
    """
    # Créer un workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"

    # Définir les largeurs de colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        'Numéro Facture',
        'Numéro Commande',
        'Livreur',
        'Statut',
        'Sous-total',
        'Frais Livraison',
        'Taxe',
        'Total'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Données
    for row, invoice in enumerate(queryset, 2):
        ws.cell(row=row, column=1).value = invoice.invoice_number
        ws.cell(row=row, column=2).value = invoice.order.order_number
        ws.cell(row=row, column=3).value = invoice.driver.user.get_full_name() or invoice.driver.user.username
        ws.cell(row=row, column=4).value = invoice.get_status_display()
        ws.cell(row=row, column=5).value = invoice.subtotal
        ws.cell(row=row, column=6).value = invoice.shipping_cost
        ws.cell(row=row, column=7).value = invoice.tax
        ws.cell(row=row, column=8).value = invoice.total

        # Appliquer les styles aux données
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format devise pour les colonnes de montants
            if col >= 5:
                cell.number_format = '#,##0.00 "TND"'

    # Ligne totale
    last_row = len(queryset) + 2
    ws.cell(row=last_row, column=1).value = "TOTAL"
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    ws.cell(row=last_row, column=8).value = f"=SUM(H2:H{last_row-1})"
    ws.cell(row=last_row, column=8).font = Font(bold=True)
    ws.cell(row=last_row, column=8).number_format = '#,##0.00 "TND"'

    # Retourner le fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


export_invoices_to_excel.short_description = "Exporter les factures sélectionnées en Excel"


def export_invoices_to_pdf(modeladmin, request, queryset):
    """
    Action pour exporter les factures sélectionnées en PDF (chaque facture avec le nouveau format)
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from django.utils.html import escape

    # Créer le PDF contenant toutes les factures
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.4*inch, bottomMargin=0.4*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)

    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1a3a52'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#1a3a52'),
        spaceAfter=4,
        alignment=TA_CENTER
    )

    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1a3a52'),
        spaceAfter=6
    )

    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4
    )

    elements = []

    # Générer chaque facture
    for idx, invoice in enumerate(queryset):
        if idx > 0:
            elements.append(PageBreak())

        # Titre
        elements.append(Paragraph("FACTURE", title_style))
        elements.append(Spacer(1, 0.15*inch))

        # Invoice number and date - centered
        invoice_info = f"<b>Numéro de Facture:</b> {escape(invoice.invoice_number)}<br/><b>Date:</b> {invoice.issued_at.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(invoice_info, info_style))
        elements.append(Spacer(1, 0.2*inch))

        # From / To section - DE is fixed, FACTURE À changes by client
        from_to_data = [
            ['DE:', 'FACTURE À:'],
            [
                "<b>HubShop</b><br/>Tunisia<br/>Tél: +216 XX XXX XXX<br/>Email: contact@hubshop.tn",
                f"<b>{escape(invoice.customer_name)}</b><br/>{escape(invoice.customer_phone)}<br/>Email: {escape(invoice.customer_email)}<br/>{escape(invoice.customer_address)}"
            ]
        ]
        from_to_table = Table(from_to_data, colWidths=[3.25*inch, 3.25*inch])
        from_to_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('FONT', (0, 1), (-1, 1), 'Helvetica', 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(from_to_table)
        elements.append(Spacer(1, 0.25*inch))

        # Order details title
        elements.append(Paragraph("Détails de la Commande:", section_title_style))

        # Order items table
        order_data = [
            ['Description', 'Quantité', 'Prix Unitaire', 'Montant'],
        ]

        # Add items
        has_items = False
        for item in invoice.order.items.all():
            has_items = True
            order_data.append([
                escape(item.product_name),
                str(item.quantity),
                f"{item.price:.2f} TND",
                f"{item.subtotal:.2f} TND"
            ])

        # If no items, add the order as a single line
        if not has_items:
            order_data.append([
                f"Commande #{escape(invoice.order.order_number)}",
                '1',
                f"{invoice.subtotal:.2f} TND",
                f"{invoice.subtotal:.2f} TND"
            ])

        order_table = Table(order_data, colWidths=[2.8*inch, 1*inch, 1.5*inch, 1.5*inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f8f8')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        elements.append(order_table)
        elements.append(Spacer(1, 0.15*inch))

        # Totals section
        totals_data = [
            ['', 'Sous-total:', f"{invoice.subtotal:.2f} TND"],
            ['', 'Frais de Livraison:', f"{invoice.shipping_cost:.2f} TND"],
            ['', 'Taxe:', f"{invoice.tax:.2f} TND"],
            ['', 'TOTAL:', f"{invoice.total:.2f} TND"],
        ]
        totals_table = Table(totals_data, colWidths=[2.8*inch, 1.8*inch, 1.6*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (1, 0), (-1, -3), 'Helvetica'),
            ('FONTNAME', (1, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (-1, -1), 9),
            ('TOPPADDING', (1, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (1, 0), (-1, -1), 4),
        ]))
        elements.append(totals_table)
        elements.append(Spacer(1, 0.2*inch))

        # Payment method and status
        payment_method = getattr(invoice.order, 'payment_method', 'online')
        method_display = 'online' if payment_method == 'online' else payment_method

        elements.append(Paragraph(f"<b>Méthode de Paiement:</b> {method_display}", normal_style))
        elements.append(Paragraph(f"<b>Statut:</b> {invoice.get_status_display()}", normal_style))

        # Footer
        footer_text = "Merci pour votre commande! / Thank you for your order!<br/>Générée le {0}".format(
            datetime.now().strftime('%d/%m/%Y à %H:%M')
        )
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER
        )
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(footer_text, footer_style))

    # Build PDF
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


export_invoices_to_pdf.short_description = "Exporter les factures sélectionnées en PDF"
