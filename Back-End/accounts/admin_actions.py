"""
Generic admin actions for exporting data to PDF and Excel formats
"""
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.utils.html import escape


def export_to_pdf(modeladmin, request, queryset):
    """
    Generic action to export queryset to PDF
    Works with any model - extracts list_display fields
    """
    # Get the model name
    model_name = queryset.model._meta.verbose_name_plural

    # Get fields to display from list_display
    if hasattr(modeladmin, 'list_display'):
        fields = modeladmin.list_display
    else:
        fields = [f.name for f in queryset.model._meta.get_fields()][:5]

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)

    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a3a52'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    # Normal style for table
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=8,
        spaceAfter=2
    )

    elements = []

    # Title
    elements.append(Paragraph(f"Export {model_name}", title_style))
    elements.append(Spacer(1, 0.2*inch))

    # Prepare table data
    table_data = []

    # Header row
    headers = []
    for field in fields:
        if hasattr(queryset.model, field):
            field_obj = queryset.model._meta.get_field(field)
            header = field_obj.verbose_name.title() if hasattr(field_obj, 'verbose_name') else field.title()
        else:
            header = field.title()
        headers.append(Paragraph(f"<b>{header}</b>", normal_style))

    table_data.append(headers)

    # Data rows
    for obj in queryset:
        row = []
        for field in fields:
            if hasattr(obj, field):
                value = getattr(obj, field)
                if callable(value):
                    value = value()
                row.append(Paragraph(str(escape(str(value)[:100])), normal_style))
            else:
                row.append(Paragraph("N/A", normal_style))
        table_data.append(row)

    # Create table
    if len(table_data) > 1:
        col_widths = [6.5*inch / len(fields)] * len(fields)
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f8f8')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ]))
        elements.append(table)

    # Build PDF
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


export_to_pdf.short_description = "Exporter en PDF"


def export_to_excel(modeladmin, request, queryset):
    """
    Generic action to export queryset to Excel
    Works with any model - extracts list_display fields
    """
    # Get the model name
    model_name = queryset.model._meta.verbose_name_plural

    # Get fields to display from list_display
    if hasattr(modeladmin, 'list_display'):
        fields = modeladmin.list_display
    else:
        fields = [f.name for f in queryset.model._meta.get_fields()][:10]

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name[:31]  # Excel sheet name limit

    # Styles
    header_fill = PatternFill(start_color="1a5490", end_color="1a5490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths
    col_width = int(72 / len(fields))
    for col_idx in range(1, len(fields) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(15, col_width)

    # Header row
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx)
        if hasattr(queryset.model, field):
            try:
                field_obj = queryset.model._meta.get_field(field)
                header = field_obj.verbose_name.title() if hasattr(field_obj, 'verbose_name') else field.title()
            except:
                header = field.title()
        else:
            header = field.title()

        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for row_idx, obj in enumerate(queryset, 2):
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if hasattr(obj, field):
                value = getattr(obj, field)
                if callable(value):
                    value = value()
                cell.value = str(value)[:100]
            else:
                cell.value = "N/A"

            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Return Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


export_to_excel.short_description = "Exporter en Excel"
